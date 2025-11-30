#!/usr/bin/env python3
"""
Модуль для конвертации заполненного JSON в Excel таблицу
С учетом уровней вложенности (1, 2, 3) в столбце E
"""

import json
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("Для работы с Excel установите: pip install openpyxl")


class JSONToExcelConverter:
    """Конвертирует заполненный JSON в форматированную Excel таблицу с уровнями вложенности"""
    
    def __init__(self):
        # Цвета для форматирования
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Серый для заголовков
        self.value_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Желтый для значений
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.header_font = Font(bold=True)
    
    def _is_parameter(self, value: Dict[str, Any]) -> bool:
        """
        Проверяет, является ли словарь параметром (имеет поля "значение", "единица" и т.д.)
        
        Args:
            value: Словарь для проверки
        
        Returns:
            True, если это параметр
        """
        if not isinstance(value, dict):
            return False
        
        # Параметр имеет поля "значение" или "единица", и все значения - примитивные типы
        has_param_fields = "значение" in value or "единица" in value
        all_values_primitive = all(
            not isinstance(v, dict) for v in value.values()
        )
        
        return has_param_fields and all_values_primitive
    
    def _parse_json_structure(self, data: Dict[str, Any], current_level: int = 1, 
                             parent_path: List[str] = None) -> List[Dict[str, Any]]:
        """
        Парсит JSON структуру и определяет уровни вложенности
        
        Args:
            data: JSON данные
            current_level: Текущий уровень вложенности (1, 2, или 3)
            parent_path: Путь к текущему элементу
        
        Returns:
            Список словарей с информацией о каждом элементе:
            {
                'name': str,
                'level': int,
                'unit': Optional[str],
                'value': Any,
                'path': List[str]
            }
        """
        if parent_path is None:
            parent_path = []
        
        items = []
        
        for key, value in data.items():
            current_path = parent_path + [key]
            
            if self._is_parameter(value):
                # Это параметр - уровень 3
                param_value = value.get("значение")
                unit = value.get("единица")
                confidence = value.get("уверенность")
                comment = value.get("комментарий")
                source = value.get("источник")
                
                items.append({
                    'name': key,
                    'level': 3,
                    'unit': unit if unit else None,
                    'value': param_value,
                    'confidence': confidence if confidence else None,
                    'comment': comment if comment else None,
                    'source': source if source else None,
                    'path': current_path
                })
            else:
                # Это секция или подсекция
                if current_level == 1:
                    # Секция верхнего уровня - уровень 1
                    items.append({
                        'name': key,
                        'level': 1,
                        'unit': None,
                        'value': None,
                        'confidence': None,
                        'comment': None,
                        'source': None,
                        'path': current_path
                    })
                    
                    # Рекурсивно обрабатываем вложенные элементы (уровень 2)
                    if isinstance(value, dict):
                        items.extend(self._parse_json_structure(value, current_level=2, parent_path=current_path))
                else:
                    # Подсекция - уровень 2
                    items.append({
                        'name': key,
                        'level': 2,
                        'unit': None,
                        'value': None,
                        'confidence': None,
                        'comment': None,
                        'source': None,
                        'path': current_path
                    })
                    
                    # Рекурсивно обрабатываем вложенные элементы
                    if isinstance(value, dict):
                        # Определяем следующий уровень
                        # Если внутри есть параметры - следующий уровень 3
                        # Если внутри только подсекции - следующий уровень 2
                        next_level = 3 if any(self._is_parameter(v) for v in value.values() if isinstance(v, dict)) else 2
                        items.extend(self._parse_json_structure(value, current_level=next_level, parent_path=current_path))
        
        return items
    
    def convert(self, json_data: Dict[str, Any], output_path: str) -> str:
        """
        Конвертирует JSON в Excel файл с уровнями вложенности
        
        Args:
            json_data: Заполненный JSON словарь
            output_path: Путь для сохранения Excel файла
        
        Returns:
            Путь к созданному файлу
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "ТЗ"
        
        # Парсим структуру JSON с определением уровней
        items = self._parse_json_structure(json_data)
        
        # Заголовки колонок (если нужны)
        # В соответствии с форматом Excel: A - название, D - единица, E - уровень
        # Столбцы B и C можно оставить пустыми или использовать для других целей
        
        # Заполняем данные
        row_num = 1
        for item in items:
            # Столбец A: название
            ws[f'A{row_num}'] = item['name']
            ws[f'A{row_num}'].border = self.border
            ws[f'A{row_num}'].alignment = Alignment(vertical='center', wrap_text=True)
            
            # Столбец B: значение (только для уровня 3)
            if item['level'] == 3:
                value = item.get('value')
                if value is not None:
                    # Форматируем значение
                    if isinstance(value, (dict, list)):
                        value_str = json.dumps(value, ensure_ascii=False)
                    else:
                        value_str = str(value)
                    ws[f'B{row_num}'] = value_str
            ws[f'B{row_num}'].border = self.border
            ws[f'B{row_num}'].alignment = Alignment(vertical='center', wrap_text=True)
            
            # Столбец C: уверенность (только для уровня 3)
            if item['level'] == 3:
                confidence = item.get('confidence')
                if confidence:
                    ws[f'C{row_num}'] = confidence
            ws[f'C{row_num}'].border = self.border
            ws[f'C{row_num}'].alignment = Alignment(vertical='center', horizontal='center')
            
            # Столбец D: единица измерения (только для уровня 3)
            if item['level'] == 3:
                unit = item.get('unit')
                if unit:
                    ws[f'D{row_num}'] = unit
            ws[f'D{row_num}'].border = self.border
            ws[f'D{row_num}'].alignment = Alignment(vertical='center', horizontal='center')
            
            # Столбец E: уровень вложенности (скрываем)
            ws[f'E{row_num}'] = item['level']
            ws[f'E{row_num}'].border = self.border
            ws[f'E{row_num}'].alignment = Alignment(vertical='center', horizontal='center')
            
            # Столбец F: комментарий (только для уровня 3)
            if item['level'] == 3:
                comment = item.get('comment')
                if comment:
                    ws[f'F{row_num}'] = comment
            ws[f'F{row_num}'].border = self.border
            ws[f'F{row_num}'].alignment = Alignment(vertical='center', wrap_text=True)
            
            # Столбец G: источник (только для уровня 3)
            if item['level'] == 3:
                source = item.get('source')
                if source:
                    ws[f'G{row_num}'] = source
            ws[f'G{row_num}'].border = self.border
            ws[f'G{row_num}'].alignment = Alignment(vertical='center', wrap_text=True)
            
            # Форматирование в зависимости от уровня
            if item['level'] == 1:
                # Секция верхнего уровня - серый фон, жирный шрифт
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    cell = ws[f'{col}{row_num}']
                    cell.fill = self.header_fill
                    cell.font = self.header_font
            elif item['level'] == 2:
                # Подсекция - серый фон, жирный шрифт
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    cell = ws[f'{col}{row_num}']
                    cell.fill = self.header_fill
                    cell.font = self.header_font
            else:
                # Параметр (уровень 3) - обычное форматирование
                # Если есть значение, добавляем желтый фон в колонку B
                if item.get('value') is not None and item.get('value') != "":
                    ws[f'B{row_num}'].fill = self.value_fill
            
            row_num += 1
        
        # Настраиваем ширину колонок
        ws.column_dimensions['A'].width = 60  # Название
        ws.column_dimensions['B'].width = 30  # Значение
        ws.column_dimensions['C'].width = 15  # Уверенность
        ws.column_dimensions['D'].width = 15  # Единица измерения
        ws.column_dimensions['E'].width = 10  # Уровень вложенности (скрыта)
        ws.column_dimensions['F'].width = 50  # Комментарий
        ws.column_dimensions['G'].width = 40  # Источник
        
        # Скрываем столбец E (уровень вложенности)
        ws.column_dimensions['E'].hidden = True
        
        # Фиксируем первую строку (если есть заголовки)
        # ws.freeze_panes = 'A2'
        
        # Сохраняем файл
        output_path_obj = Path(output_path)
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path_obj))
        
        return str(output_path_obj)


if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 3:
        print("Использование: python json_to_excel.py <путь_к_json_файлу> <путь_для_excel_файла>")
        sys.exit(1)
    
    json_file = sys.argv[1]
    excel_file = sys.argv[2]
    
    try:
        # Загружаем JSON
        with open(json_file, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        # Конвертируем
        converter = JSONToExcelConverter()
        output_path = converter.convert(json_data, excel_file)
        
        print(f"✅ JSON успешно конвертирован в Excel: {output_path}")
        
    except Exception as e:
        print(f"❌ Ошибка: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)

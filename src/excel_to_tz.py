#!/usr/bin/env python3
"""
Модуль для конвертации Excel шаблона в TZ.json
Создает/обновляет TZ.json на основе структуры Excel файла с уровнями вложенности в столбце E
"""

import json
import openpyxl
from pathlib import Path
from typing import Dict, Any, Optional, List, Tuple


class ExcelToTZConverter:
    """Конвертер Excel шаблона в TZ.json"""
    
    def __init__(self, excel_path: str, tz_json_path: str = "TZ.json"):
        """
        Инициализация конвертера
        
        Args:
            excel_path: Путь к Excel файлу
            tz_json_path: Путь к файлу TZ.json для обновления
        """
        self.excel_path = Path(excel_path)
        project_root = Path(__file__).parent.parent
        self.tz_json_path = project_root / tz_json_path if not Path(tz_json_path).is_absolute() else Path(tz_json_path)
        
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel файл не найден: {excel_path}")
    
    def load_tz_json(self) -> Dict[str, Any]:
        """Загружает существующий TZ.json или создает пустой"""
        if self.tz_json_path.exists():
            with open(self.tz_json_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    
    def save_tz_json(self, tz_data: Dict[str, Any]):
        """Сохраняет обновленный TZ.json"""
        self.tz_json_path.parent.mkdir(parents=True, exist_ok=True)
        with open(self.tz_json_path, 'w', encoding='utf-8') as f:
            json.dump(tz_data, f, ensure_ascii=False, indent=2)
    
    def parse_excel_structure(self) -> List[Dict[str, Any]]:
        """
        Парсит Excel файл с учетом уровней вложенности из столбца E
        
        Returns:
            Список словарей с информацией о каждом элементе:
            {
                'name': str,  # Из столбца A
                'level': int,  # Из столбца E (1, 2, или 3)
                'unit': Optional[str],  # Из столбца D (только для уровня 3)
                'match': Optional[List[str]],  # Из столбца F (только для уровня 3)
                'row': int,
                'path': List[str]  # Полный путь в иерархии
            }
        """
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        
        items = []
        hierarchy_stack = []  # Стек для отслеживания текущей иерархии
        
        # Читаем все строки
        for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
            col_a = str(row[0]).strip() if row[0] else ''  # Название
            col_d = str(row[3]).strip() if len(row) > 3 and row[3] else ''  # Единица измерения
            col_e = str(row[4]).strip() if len(row) > 4 and row[4] else ''  # Уровень вложенности
            col_f = str(row[5]).strip() if len(row) > 5 and row[5] else ''  # Глоссарий (match)
            
            if not col_a:
                continue
            
            # Парсим уровень вложенности
            try:
                level = int(col_e) if col_e else None
            except (ValueError, TypeError):
                level = None
            
            if level is None:
                # Если уровень не указан, пропускаем строку
                continue
            
            if level == 1:
                # Секция верхнего уровня - сбрасываем стек и добавляем секцию
                hierarchy_stack = [col_a]
                items.append({
                    'name': col_a,
                    'level': 1,
                    'type': 'section',
                    'unit': None,
                    'match': None,
                    'row': row_num,
                    'path': [col_a]
                })
            
            elif level == 2:
                # Подсекция - находим родителя (последняя строка с уровнем < 2, т.е. уровень 1)
                # Обновляем стек: оставляем только секцию (уровень 1) и добавляем подсекцию
                if hierarchy_stack:
                    # Оставляем только первый элемент (секция уровня 1)
                    hierarchy_stack = [hierarchy_stack[0], col_a]
                else:
                    # Если нет секции, создаем пустую структуру
                    hierarchy_stack = [col_a]
                
                items.append({
                    'name': col_a,
                    'level': 2,
                    'type': 'subsection',
                    'unit': None,
                    'match': None,
                    'row': row_num,
                    'path': hierarchy_stack.copy()
                })
            
            elif level == 3:
                # Параметр - находим родителя (последняя строка с уровнем < 3, т.е. уровень 1 или 2)
                # Используем текущий стек иерархии
                if not hierarchy_stack:
                    # Если нет родителя, пропускаем
                    continue
                
                # Единица измерения берется из столбца D
                unit = col_d if col_d else None
                
                # Парсим глоссарий из столбца F
                match = self._parse_glossary(col_f)
                
                items.append({
                    'name': col_a,
                    'level': 3,
                    'type': 'parameter',
                    'unit': unit,
                    'match': match,
                    'row': row_num,
                    'path': hierarchy_stack + [col_a]
                })
        
        return items
    
    def _parse_glossary(self, glossary_text: str) -> Optional[List[str]]:
        """
        Парсит текст глоссария из столбца F в список строк
        
        Args:
            glossary_text: Текст из столбца F (может содержать несколько значений через разделители)
        
        Returns:
            Список строк или None, если пусто
        """
        if not glossary_text or not glossary_text.strip():
            return None
        
        # Разделяем по различным разделителям: запятая, точка с запятой, перенос строки
        # Сначала пробуем разделить по переносу строки
        if '\n' in glossary_text:
            items = [item.strip() for item in glossary_text.split('\n') if item.strip()]
        # Затем по точке с запятой
        elif ';' in glossary_text:
            items = [item.strip() for item in glossary_text.split(';') if item.strip()]
        # Затем по запятой
        elif ',' in glossary_text:
            items = [item.strip() for item in glossary_text.split(',') if item.strip()]
        else:
            # Если нет разделителей, возвращаем как одну строку
            items = [glossary_text.strip()]
        
        # Убираем пустые элементы и возвращаем список
        items = [item for item in items if item]
        return items if items else None
    
    def build_tz_structure(self, excel_items: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Строит структуру TZ.json на основе элементов из Excel
        
        Args:
            excel_items: Список элементов из Excel
        
        Returns:
            Словарь с полной структурой TZ.json
        """
        tz_structure = {}
        
        for item in excel_items:
            if item['type'] == 'parameter':
                # Строим путь к параметру
                path = item['path']
                
                # Начинаем с корня
                current = tz_structure
                
                # Проходим по пути, создавая структуру
                for i, key in enumerate(path[:-1]):
                    if key not in current:
                        current[key] = {}
                    current = current[key]
                
                # Добавляем параметр с полной структурой
                param_name = path[-1]
                current[param_name] = {
                    "значение": None,
                    "единица": item['unit'] if item['unit'] else None,
                    "источник": None,
                    "уверенность": None,
                    "комментарий": None
                }
        
        return tz_structure
    
    def build_glossary(self, excel_items: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Строит структуру glossary.json с той же вложенностью, что и TZ.json
        
        Args:
            excel_items: Список элементов из Excel
        
        Returns:
            Словарь с глоссарием, повторяющий структуру TZ.json:
            {
                "Секция": {
                    "Подсекция": {
                        "Параметр": {
                            "match": ["вариант1", "вариант2", ...] или null,
                            "unit": "мм" или null
                        }
                    }
                }
            }
        """
        glossary = {}
        
        for item in excel_items:
            if item['type'] == 'parameter' and item['level'] == 3:
                # Строим путь к параметру (как в build_tz_structure)
                path = item['path']
                
                # Начинаем с корня
                current = glossary
                
                # Проходим по пути, создавая структуру
                for i, key in enumerate(path[:-1]):
                    if key not in current:
                        current[key] = {}
                    current = current[key]
                
                # Добавляем параметр с match и unit
                param_name = path[-1]
                current[param_name] = {
                    "match": item.get('match'),
                    "unit": item.get('unit')
                }
        
        return glossary
    
    def save_glossary(self, glossary: Dict[str, Any], glossary_path: Optional[str] = None):
        """
        Сохраняет glossary.json
        
        Args:
            glossary: Словарь с глоссарием
            glossary_path: Путь к файлу glossary.json (если None, используется корень проекта)
        """
        if glossary_path is None:
            project_root = Path(__file__).parent.parent
            glossary_path = project_root / "glossary.json"
        else:
            glossary_path = Path(glossary_path)
        
        glossary_path.parent.mkdir(parents=True, exist_ok=True)
        with open(glossary_path, 'w', encoding='utf-8') as f:
            json.dump(glossary, f, ensure_ascii=False, indent=2)
    
    def merge_with_existing_tz(self, new_structure: Dict[str, Any], existing_tz: Dict[str, Any]) -> Dict[str, Any]:
        """
        Объединяет новую структуру из Excel с существующим TZ.json
        Сохраняет значения из существующего TZ.json, если они есть
        
        Args:
            new_structure: Новая структура из Excel
            existing_tz: Существующий TZ.json
        
        Returns:
            Объединенная структура
        """
        def merge_dict(new: Dict, existing: Dict) -> Dict:
            """Рекурсивно объединяет словари"""
            result = new.copy()
            
            for key, value in existing.items():
                if key in result:
                    if isinstance(value, dict) and isinstance(result[key], dict):
                        # Если это словари, проверяем, является ли это параметром
                        if "значение" in value or "единица" in value:
                            # Это параметр - сохраняем значения из существующего, но обновляем единицу из нового
                            result[key] = value.copy()
                            if "единица" in new[key] and new[key]["единица"]:
                                result[key]["единица"] = new[key]["единица"]
                        else:
                            # Это подсекция - рекурсивно объединяем
                            result[key] = merge_dict(result[key], value)
                    elif isinstance(value, dict) and "значение" in value:
                        # Существующий параметр - сохраняем его
                        result[key] = value.copy()
                        # Обновляем единицу из нового, если есть
                        if key in new and isinstance(new[key], dict) and "единица" in new[key]:
                            if new[key]["единица"]:
                                result[key]["единица"] = new[key]["единица"]
                else:
                    # Ключ есть только в существующем - сохраняем его
                    result[key] = value
            
            return result
        
        return merge_dict(new_structure, existing_tz)
    
    def convert(self, preserve_existing_values: bool = True, dry_run: bool = False) -> Dict[str, Any]:
        """
        Основной метод конвертации
        
        Args:
            preserve_existing_values: Если True, сохраняет значения из существующего TZ.json
            dry_run: Если True, не сохраняет изменения
        
        Returns:
            Статистика конвертации
        """
        # Парсим Excel
        excel_items = self.parse_excel_structure()
        
        # Строим новую структуру TZ.json
        new_structure = self.build_tz_structure(excel_items)
        
        # Строим глоссарий
        glossary = self.build_glossary(excel_items)
        
        # Загружаем существующий TZ.json
        existing_tz = self.load_tz_json()
        
        # Объединяем
        if preserve_existing_values and existing_tz:
            final_structure = self.merge_with_existing_tz(new_structure, existing_tz)
        else:
            final_structure = new_structure
        
        # Подсчитываем статистику
        def count_parameters(data: Dict, count: List[int] = None) -> int:
            if count is None:
                count = [0]
            for key, value in data.items():
                if isinstance(value, dict):
                    # Проверяем, является ли это параметром
                    has_param_fields = "значение" in value or "единица" in value
                    all_values_are_primitive = all(
                        not isinstance(v, dict) for v in value.values()
                    )
                    
                    if has_param_fields and all_values_are_primitive:
                        count[0] += 1
                    else:
                        # Это подсекция - рекурсивно считаем параметры в ней
                        count_parameters(value, count)
            return count[0]
        
        def count_glossary_parameters(data: Dict, count: List[int] = None) -> int:
            """Подсчитывает параметры в glossary (поля: match, unit)"""
            if count is None:
                count = [0]
            for key, value in data.items():
                if isinstance(value, dict):
                    # Проверяем, является ли это параметром glossary
                    has_glossary_fields = "match" in value or "unit" in value
                    all_values_are_primitive = all(
                        not isinstance(v, dict) for v in value.values()
                    )
                    
                    if has_glossary_fields and all_values_are_primitive:
                        count[0] += 1
                    else:
                        # Это подсекция - рекурсивно считаем параметры в ней
                        count_glossary_parameters(value, count)
            return count[0]
        
        total_params = count_parameters(final_structure)
        excel_params = len([item for item in excel_items if item['type'] == 'parameter'])
        # Подсчитываем параметры в glossary (рекурсивно)
        glossary_params = count_glossary_parameters(glossary)
        
        stats = {
            'total_parameters_in_excel': excel_params,
            'total_parameters_in_tz': total_params,
            'total_parameters_in_glossary': glossary_params,
            'sections': len([item for item in excel_items if item['type'] == 'section']),
            'subsections': len([item for item in excel_items if item['type'] == 'subsection']),
        }
        
        # Сохраняем, если не dry_run
        if not dry_run:
            self.save_tz_json(final_structure)
            self.save_glossary(glossary)
        
        return stats


if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 2:
        print("Использование: python excel_to_tz.py <путь_к_excel_файлу> [--dry-run] [--no-preserve]")
        print("  --dry-run: не сохранять изменения, только показать статистику")
        print("  --no-preserve: не сохранять существующие значения из TZ.json")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    dry_run = '--dry-run' in sys.argv
    preserve = '--no-preserve' not in sys.argv
    
    try:
        converter = ExcelToTZConverter(excel_file)
        stats = converter.convert(preserve_existing_values=preserve, dry_run=dry_run)
        
        print("=" * 80)
        print("📊 Результаты конвертации Excel → TZ.json и glossary.json")
        print("=" * 80)
        print(f"Секций верхнего уровня: {stats['sections']}")
        print(f"Подсекций: {stats['subsections']}")
        print(f"Параметров в Excel: {stats['total_parameters_in_excel']}")
        print(f"Параметров в TZ.json: {stats['total_parameters_in_tz']}")
        print(f"Параметров в glossary.json: {stats['total_parameters_in_glossary']}")
        
        if not dry_run:
            print(f"\n💾 Структура TZ.json сохранена в {converter.tz_json_path}")
            glossary_path = converter.tz_json_path.parent / "glossary.json"
            print(f"💾 Глоссарий сохранен в {glossary_path}")
        else:
            print(f"\n🔍 Режим dry-run: изменения НЕ сохранены")
        
    except Exception as e:
        print(f"❌ Ошибка: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)

#!/usr/bin/env python3
"""
Модуль для добавления CSV данных в Excel файл как отдельные листы
"""

import csv
import io
import re
from pathlib import Path
from typing import Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class CSVToExcelAppender:
    """Добавляет CSV данные в существующий Excel файл как новый лист"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.header_font = Font(bold=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def parse_csv_from_text(self, text: str) -> str:
        """
        Извлекает CSV из текста ответа AI
        
        Args:
            text: Текст ответа от AI (может содержать markdown, пояснения и т.д.)
        
        Returns:
            Чистый CSV текст
        """
        if not text or not text.strip():
            return ""
        
        # Убираем markdown блоки кода если есть
        if '```' in text:
            # Ищем блок кода (может быть ```csv, ``` или просто ```)
            # Паттерн ищет ``` опционально с языком, затем содержимое, затем закрывающий ```
            pattern = r'```(?:csv|CSV)?\s*\n?(.*?)```'
            match = re.search(pattern, text, re.DOTALL)
            if match:
                text = match.group(1).strip()
            else:
                # Пробуем найти просто между ```
                parts = text.split('```')
                if len(parts) >= 3:
                    # Берем содержимое между первыми ```
                    potential_text = parts[1].strip()
                    # Убираем префикс языка если есть
                    if potential_text.lower().startswith('csv'):
                        potential_text = potential_text[3:].strip()
                    # Проверяем что это похоже на CSV (содержит запятые)
                    if ',' in potential_text:
                        text = potential_text
        
        # Ищем таблицу (строки с запятыми или табуляцией)
        lines = text.split('\n')
        csv_lines = []
        in_table = False
        
        for line in lines:
            line = line.strip()
            if not line:
                # Пропускаем пустые строки, но если мы в таблице, продолжаем
                if in_table:
                    continue
                else:
                    continue
            
            # Если строка содержит запятые - это CSV
            if ',' in line:
                # Убираем markdown форматирование таблиц (|)
                line = line.replace('|', '').strip()
                # Убираем лишние пробелы вокруг запятых
                line = re.sub(r'\s*,\s*', ',', line)
                csv_lines.append(line)
                in_table = True
            elif in_table and (line.startswith('|') or ',' in line):
                # Продолжение таблицы
                line = line.replace('|', '').strip()
                line = re.sub(r'\s*,\s*', ',', line)
                csv_lines.append(line)
            elif in_table:
                # Конец таблицы - проверяем, не пустая ли строка
                if line and not line.startswith('#'):
                    # Возможно, это продолжение данных без запятых
                    continue
                else:
                    break
        
        if csv_lines:
            return '\n'.join(csv_lines)
        
        # Если не нашли CSV, пробуем парсить JSON-подобные объекты
        json_csv = self._parse_json_like_to_csv(text)
        if json_csv:
            return json_csv
        
        # Если не нашли таблицу, возвращаем весь текст (может быть уже CSV)
        return text.strip()
    
    def _parse_json_like_to_csv(self, text: str) -> str:
        """
        Парсит JSON-подобные объекты и конвертирует их в CSV
        
        Формат входных данных:
        name: "Значение"
        type: "тип"
        quantity: 3
        ...
        или
        ref: "66.5"
        name: "Люнет"
        type: "люнет"
        ...
        
        Args:
            text: Текст с JSON-подобными объектами
        
        Returns:
            CSV строка или пустая строка если не найдено
        """
        if not text or not text.strip():
            return ""
        
        # Ищем объекты в формате key: value
        objects = []
        current_object = {}
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip()
            
            # Пропускаем пустые строки и закрывающие скобки
            if not line:
                continue
            
            if line == '}' or line == ']':
                if current_object:
                    objects.append(current_object)
                    current_object = {}
                continue
            
            # Пропускаем открывающие скобки и комментарии
            if line == '{' or line == '[' or line.startswith('//') or line.startswith('#'):
                continue
            
            # Парсим key: value
            # Поддерживаем форматы:
            # key: "value"
            # key: value
            # key: null
            # Также поддерживаем ключи с пробелами и кириллицей: "ключ": "значение" или ключ: "значение"
            # Паттерн: ключ (может быть в кавычках или без) : значение
            match = re.match(r'["\']?([^:]+?)["\']?\s*:\s*(.+)$', line)
            if match:
                key = match.group(1).strip()
                value = match.group(2).strip()
                
                # Обрабатываем значение
                if value.lower() == 'null':
                    value = ''
                elif value.startswith('"') and value.endswith('"'):
                    # Убираем кавычки
                    value = value[1:-1]
                elif value.startswith("'") and value.endswith("'"):
                    # Убираем одинарные кавычки
                    value = value[1:-1]
                
                current_object[key] = value
        
        # Добавляем последний объект если есть
        if current_object:
            objects.append(current_object)
        
        if not objects:
            return ""
        
        # Определяем заголовки на основе всех объектов (объединение всех ключей)
        all_keys = set()
        for obj in objects:
            all_keys.update(obj.keys())
        
        # Определяем тип данных по ключам (для правильного маппинга)
        # Услуги обычно имеют description и type с монтаж/обучение
        # Оснастка имеет type="люнет/патрон" и может иметь "диапазон зажима"
        # ЗИП имеет quantity и unit или категорию
        has_service_keywords = any('услуга' in k.lower() or 'service' in k.lower() for k in all_keys)
        if not has_service_keywords and 'description' in all_keys:
            # Проверяем значения type в объектах
            for obj in objects:
                obj_type = str(obj.get('type', '')).lower()
                if 'монтаж' in obj_type or 'обучение' in obj_type or 'пусконаладка' in obj_type or \
                   'инжиниринг' in obj_type or 'программирование' in obj_type:
                    has_service_keywords = True
                    break
            # Если есть description и нет специфичных для оснастки ключей - это услуги
            if not has_service_keywords and 'description' in all_keys and \
               not any('диапазон' in k.lower() or 'зажима' in k.lower() for k in all_keys) and \
               not any('тип' in k.lower() and any('люнет' in str(obj.get('type', '')).lower() or 
                                                   'патрон' in str(obj.get('type', '')).lower() 
                                                   for obj in objects) for k in all_keys):
                has_service_keywords = True
        
        # Для объединенного промпта инструмент+оснастка проверяем наличие типа и инструмента/оснастки
        has_tooling_keywords = any('тип' in k.lower() or 'type' in k.lower() for k in all_keys) and \
                              (any('диапазон' in k.lower() or 'зажима' in k.lower() for k in all_keys) or
                               any('инструмент' in str(obj.get('type', '')).lower() or 
                                   'оснастка' in str(obj.get('type', '')).lower() or
                                   'патрон' in str(obj.get('type', '')).lower() or
                                   'люнет' in str(obj.get('type', '')).lower()
                                   for obj in objects))
        
        has_spare_parts_keywords = any('категория' in k.lower() or 'category' in k.lower() for k in all_keys)
        if not has_spare_parts_keywords:
            # Проверяем наличие unit с % или специфичные для ЗИП паттерны
            for obj in objects:
                unit_val = str(obj.get('unit', '')).lower()
                if '%' in unit_val or 'стоимости' in unit_val:
                    has_spare_parts_keywords = True
                    break
        
        # Маппинг ключей на стандартные названия колонок
        key_mapping = {
            'name': 'Услуга' if has_service_keywords else 'Наименование',
            'ref': 'Источник',
            'type': 'Тип',
            'quantity': 'Количество',
            'unit': 'Единица',
            'article': 'Артикул',
            'description': 'Описание/условия',
            'описание': 'Описание/условия',
            'источник': 'Источник',
            'уверенность': 'Уверенность',
            'примечание': 'Примечание',
            'категория': 'Категория',
            'условия': 'Описание/условия',
            'наличие': 'Наличие',
            'диапазон зажима': 'Примечание',  # Для оснастки это примечание
            'услуга': 'Услуга'
        }
        
        # Определяем порядок заголовков в зависимости от типа данных
        if has_service_keywords:
            # Для услуг: Услуга, Описание/условия, Источник
            priority_order = ['Услуга', 'Описание/условия', 'Источник', 'Тип']
            # Переименовываем name в Услуга
            key_mapping['name'] = 'Услуга'
            # Если есть ref, маппим его в Источник
            if 'ref' in all_keys:
                key_mapping['ref'] = 'Источник'
        elif has_tooling_keywords:
            # Для оснастки: Наименование, Количество, Тип, Примечание, Источник, Уверенность
            priority_order = ['Наименование', 'Количество', 'Тип', 'Примечание', 'Источник', 'Уверенность', 'Единица']
            # Маппим "диапазон зажима" в Примечание
            for key in all_keys:
                if 'диапазон' in key.lower() or 'зажима' in key.lower():
                    key_mapping[key] = 'Примечание'
        elif has_spare_parts_keywords:
            # Для ЗИП: Наименование, Количество, Категория, Источник, Уверенность
            priority_order = ['Наименование', 'Количество', 'Категория', 'Источник', 'Уверенность', 'Единица']
        else:
            # По умолчанию: Наименование, Количество, Тип, Описание, Источник
            priority_order = ['Наименование', 'Количество', 'Тип', 'Описание/условия', 'Источник', 'Уверенность', 'Единица']
        
        # Собираем заголовки
        mapped_headers = set()
        for key in all_keys:
            mapped_key = key_mapping.get(key, key)
            mapped_headers.add(mapped_key)
        
        # Сортируем по приоритету
        headers = []
        for priority_header in priority_order:
            if priority_header in mapped_headers:
                headers.append(priority_header)
                mapped_headers.remove(priority_header)
        
        # Добавляем остальные заголовки
        headers.extend(sorted(mapped_headers))
        
        # Обратный маппинг для поиска значений
        reverse_mapping = {}
        for key in all_keys:
            mapped_key = key_mapping.get(key, key)
            if mapped_key not in reverse_mapping:
                reverse_mapping[mapped_key] = []
            reverse_mapping[mapped_key].append(key)
        
        if not headers:
            return ""
        
        # Формируем CSV
        csv_lines = []
        # Заголовок
        csv_lines.append(','.join(headers))
        
        # Данные
        for obj in objects:
            row = []
            for header in headers:
                # Ищем значение по маппингу
                value = ''
                if header in reverse_mapping:
                    for original_key in reverse_mapping[header]:
                        if original_key in obj:
                            value = str(obj[original_key])
                            break
                else:
                    # Пробуем напрямую
                    value = str(obj.get(header, ''))
                
                # Экранируем запятые и кавычки в значениях
                if ',' in value or '"' in value or '\n' in value:
                    value = '"' + value.replace('"', '""') + '"'
                row.append(value)
            csv_lines.append(','.join(row))
        
        return '\n'.join(csv_lines)
    
    def add_csv_sheet(self, excel_path: str, csv_text: str, sheet_name: str) -> str:
        """
        Добавляет CSV данные в Excel файл как новый лист
        
        Args:
            excel_path: Путь к существующему Excel файлу
            csv_text: CSV текст (строка)
            sheet_name: Имя нового листа
        
        Returns:
            Путь к обновленному Excel файлу
        """
        if not csv_text or not csv_text.strip():
            # Если CSV пустой, создаем лист с сообщением
            wb = load_workbook(excel_path)
            ws = wb.create_sheet(title=sheet_name)
            ws['A1'] = "Данные не найдены"
            ws['A1'].font = Font(italic=True)
            wb.save(excel_path)
            return excel_path
        
        # Загружаем существующую книгу
        wb = load_workbook(excel_path)
        
        # Проверяем, не существует ли уже лист с таким именем
        if sheet_name in wb.sheetnames:
            # Удаляем старый лист
            wb.remove(wb[sheet_name])
        
        # Создаем новый лист
        ws = wb.create_sheet(title=sheet_name)
        
        # Парсим CSV
        try:
            csv_reader = csv.reader(io.StringIO(csv_text))
            rows = list(csv_reader)
        except Exception as e:
            # Если не удалось распарсить, создаем лист с ошибкой
            ws['A1'] = f"Ошибка парсинга CSV: {str(e)}"
            ws['A2'] = "Исходный текст:"
            ws['A3'] = csv_text[:1000]  # Первые 1000 символов
            wb.save(excel_path)
            return excel_path
        
        if not rows:
            # Если CSV пустой, добавляем заголовок
            ws['A1'] = "Данные не найдены"
            ws['A1'].font = Font(italic=True)
            wb.save(excel_path)
            return excel_path
        
        # Записываем данные
        for row_idx, row in enumerate(rows, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Очищаем значение от лишних пробелов
                cell.value = value.strip() if value and value.strip() else ""
                
                # Форматируем заголовок (первая строка)
                if row_idx == 1:
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                cell.border = self.border
        
        # Настраиваем ширину колонок
        if rows:
            num_cols = len(rows[0])
            for col_idx in range(1, num_cols + 1):
                col_letter = get_column_letter(col_idx)
                # Автоматическая ширина на основе содержимого
                max_length = 0
                for row in rows:
                    if col_idx <= len(row):
                        cell_value = str(row[col_idx - 1]) if row[col_idx - 1] else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                
                # Устанавливаем ширину с небольшим запасом
                ws.column_dimensions[col_letter].width = min(max(max_length + 2, 15), 50)
        
        # Сохраняем
        wb.save(excel_path)
        return excel_path

